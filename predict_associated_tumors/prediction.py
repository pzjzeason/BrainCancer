import json
from openpyxl import load_workbook
from pandas import DataFrame
import matplotlib.pyplot as plt

data = {}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_with_associatedTumours2.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        values = {}
        line = json.loads(each)

        for value in line:
            if value != 'id':
                values[value]=line[value]

        data[line['id']] = values

# 预测集
wb = load_workbook('/Users/燚/study/大三下/信息系统开发与设计/系统开发/prediction_set.xlsx')
sheet = wb.active
# 预测集id保存在列表training_set
prediction_set = []
for row in range(2, sheet.max_row+1):
    id_cell = sheet.cell(row=row, column=1)  # 读取训练集中肿瘤id
    id_text = id_cell.value
    prediction_set.append(int(id_text))

# 创建数据结构 DataFrame矩阵score_dataFrame
score_dataFrame = DataFrame(columns=prediction_set, index=prediction_set)
for id in data:
    if int(id) in prediction_set:
        for each in data[id]:  # 确保两两计算的肿瘤均属于训练集
            if int(each) in prediction_set:
                score_dataFrame.loc[int(id),int(each)]=[]
                score_dataFrame.loc[int(id),int(each)].append(data[id][each]['brainStructure_score'])
                score_dataFrame.loc[int(id),int(each)].append(data[id][each]['symptoms_score'])
                score_dataFrame.loc[int(id), int(each)].append(data[id][each]['image_score'])
                score_dataFrame.loc[int(id), int(each)].append(data[id][each]['isAssociated'])

wb = load_workbook('/Users/燚/study/大三下/信息系统开发与设计/系统开发/prediction.xlsx')
sheet = wb.active  # 获得当前正在显示的sheet

count = 0  # 预测相似计数器
for row_id in prediction_set:
    for col_id in prediction_set:
        if col_id > row_id:
            if score_dataFrame.loc[row_id, col_id][3] == 0: # 无相似记录
                score = score_dataFrame.loc[row_id, col_id][2]*1  # 总分
                if score >= 0.27:
                    count +=1
                    sheet.cell(row=count+1, column=1).value = row_id
                    sheet.cell(row=count+1, column=2).value = col_id
wb.save('/Users/燚/study/大三下/信息系统开发与设计/系统开发/prediction.xlsx')

