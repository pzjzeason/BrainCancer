import json

from openpyxl import load_workbook
from openpyxl.cell import Cell

wb = load_workbook('抽取数据(脑区、症状、易感人群、概率)2.0.xlsx')

sheet = wb.active

# 用于存储数据的数组
list1 = []
list2 = []
list3 = []

for row in range(2, sheet.max_row + 1):
    id_cell = sheet.cell(row=row, column=1)
    id_text = id_cell.value

   # 将表中第一列的id编号写入list1数组中
    list1.append(id_text)

   # 将表中第2列的脑区写入list2数组中
    nq_text0 = ""
    nq_cell = sheet.cell(row=row, column=2)
    nq_text0 = nq_cell.value
    if nq_text0:
        nq_text = nq_text0
    else:
        nq_text = None
    list2.append(nq_text)

    # 将表中第3列的临床病症写入list3数组中
    bz_text0 = ""
    bz_cell = sheet.cell(row=row, column=3)
    bz_text0 = bz_cell.value
    if bz_text0:
        bz_text = bz_text0
    else:
        bz_text = None
    list3.append(bz_text)


# 遍历所有肿瘤，计算两两病症之间易发脑区相似度分数
# 首先遍历所有肿瘤，为每个肿瘤构建列表，用于存储其与其他肿瘤的脑区相似度分数：

for i in range(0, len(list1)):
    data = {}
    data1 = {}

    score1 = []   # 用于存储脑区分数的列表

    score2 = []   # 用于存储临床表现分数的列表

    list_A = list2[i]    # 该肿瘤的易发脑区

    list_a = list3[i]    # 该肿瘤的临床病症


# 两两遍历所有肿瘤，计算分数，将得到的值加入列表score1中：

# 计算易发脑区相似度分数

    for j in range(0, len(list1)):
        list_B = list2[j]   # 其他某一肿瘤的易发脑区

        # 计算临床病症相似度分数
        list_b = list3[j]  # 其他某一肿瘤的临床表现

        count1 = 0
        count2 = 0

         # 若两个文本都不为空
        if list_A is not None and list_B is not None:

             # 如果是anywhere，脑区相似度分数则为1
            if list_B == 'Anywhere'or list_A == 'Anywhere':
                score1.append(1)
            else:
                for list_1_element in list(list_A.split(',')):
                    for list_2_element in list(list_B.split(',')):
                        if list_2_element == list_1_element:
                            count1 += 1
                            print(list_2_element)

            # 如果存在相同的易发脑区，则该分数为1，其余情况均为0
            if 0 < count1:
                score1.append(1)
            else:
                score1.append(0)
        else:
            score1.append(0)

            # 若不为空，则比较list_a与list_b之间相同症状的个数：
        if list_a is not None and list_b is not None:
            for list_1_element in list(list_a.split(',')):
                for list_2_element in list(list_b.split(',')):
                    if list_2_element == list_1_element:
                        count2 += 1      # 有相同的症状，则加一
            if 0 < count2:
                score2.append(count2)    # 将count2值加入score2中
            else:
                score2.append(0)
        else:
            score2.append(0)

        data1[str(j + 1)] = {'brainstucture': score1[j], 'symptoms': score2[j]}
        data = dict(data, **data1)
        data["id"] = str(i + 1)

    with open("D:/Python/score.json", 'a', encoding='utf-8') as json_file:
        json_file.write(json.dumps(data, ensure_ascii=False) + '\n')