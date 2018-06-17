import json
from openpyxl import load_workbook
import pandas as pd
from pandas import Series,DataFrame
import matplotlib.pyplot as plt
import numpy as np
from mpl_toolkits.mplot3d import Axes3D

data = {}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_with_associatedTumours2.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        values = {}
        line = json.loads(each)

        for value in line:
            if value != 'id':
                values[value]=line[value]

        data[line['id']] = values

# 训练集
wb = load_workbook('/Users/燚/study/大三下/信息系统开发与设计/系统开发/training_set.xlsx')
sheet = wb.active

# 训练集id保存在列表training_set
training_set = []
for row in range(2, sheet.max_row+1):
    id_cell = sheet.cell(row=row, column=1)  # 读取训练集中肿瘤id
    id_text = id_cell.value
    training_set.append(int(id_text))

# 创建数据结构 DataFrame矩阵score_dataFrame
score_dataFrame = DataFrame(columns=training_set, index=training_set)
for id in data:
    if int(id) in training_set:
        for each in data[id]:  # 确保两两计算的肿瘤均属于训练集
            if int(each) in training_set:
                score_dataFrame.loc[int(id),int(each)]=[]
                score_dataFrame.loc[int(id),int(each)].append(data[id][each]['brainStructure_score'])
                score_dataFrame.loc[int(id),int(each)].append(data[id][each]['symptoms_score'])
                score_dataFrame.loc[int(id), int(each)].append(data[id][each]['image_score'])
                score_dataFrame.loc[int(id), int(each)].append(data[id][each]['isAssociated'])

x_data_sim = []  # 相似集数据
y_data_sim = []
x_data = []  # 不相似集数据
y_data = []
max_score_presentation=0
for row in training_set:
    for col in training_set:
        if col > row:
            if score_dataFrame.loc[row, col][2]<0.3:
                if score_dataFrame.loc[row, col][3] == 1:
                    x_data_sim.append(score_dataFrame.loc[row, col][2] * 1)  # 总分计算
                    y_data_sim.append(score_dataFrame.loc[row, col][2] * 1)
                    if score_dataFrame.loc[row, col][2] > max_score_presentation:
                        max_score_presentation = score_dataFrame.loc[row, col][2]
                else:
                    x_data.append(score_dataFrame.loc[row, col][2] * 1)
                    y_data.append(score_dataFrame.loc[row, col][2] * 1)
print(max_score_presentation)

plt.scatter(x_data, y_data, marker='x', c='black')  # 绘制数据点
plt.scatter(x_data_sim, y_data_sim, marker='o', c='r')

# 绘制x，y轴的说明
plt.xlabel('score_presentation')
plt.ylabel('score')
plt.title('Mark for Training Set')
plt.show()  # 显示


# 查准率与召回率计算
threshold_dic = DataFrame(columns=['precision','recall'])  # 结果存储表
for threshold_test in np.arange(0,max_score_presentation,0.01):  # 以0.01的递进测试阈值
    precision_record = 0
    precision_all = 0
    recall_record = 0
    recall_all = 0
    for row in training_set:
        for col in training_set:
            if col > row:
                if score_dataFrame.loc[row, col][2] >= threshold_test:
                    precision_all += 1  # 查准
                    if score_dataFrame.loc[row, col][3] == 1:
                        precision_record += 1
                        recall_record += 1  # 查全
                        recall_all += 1
                else:
                    if score_dataFrame.loc[row, col][3] == 1:
                        recall_all += 1
    if precision_all != 0:
        precision = precision_record/precision_all
    if recall_all != 0:
        recall = recall_record/recall_all
    threshold_dic.loc[threshold_test] = {'precision': precision, 'recall': recall}
print(threshold_dic)

p_data = []
r_data = []
for index in threshold_dic.index:
    p_data.append(threshold_dic.loc[index,'precision'])
    r_data.append(threshold_dic.loc[index,'recall'])

plt.plot(r_data, p_data, 'o-')  # 绘制数据点
ax = plt.gca()
for index in threshold_dic.index:  # 添加节点阈值标签
    ax.text(threshold_dic.loc[index,'recall'],threshold_dic.loc[index,'precision'],index,ha='center',fontsize=7)

# 绘制x，y轴的说明
plt.xlabel('recall')
plt.ylabel('precision')
plt.title('P-R Curve')
plt.show()  # 显示



