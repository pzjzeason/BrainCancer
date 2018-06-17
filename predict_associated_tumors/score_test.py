import json
from openpyxl import load_workbook
from pandas import DataFrame
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

data = {}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_with_associatedTumours2.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        values = {}
        line = json.loads(each)

        for value in line:
            if value != 'id':
                values[value]=line[value]

        data[line['id']] = values

# 训练集
wb = load_workbook('/Users/燚/study/大三下/信息系统开发与设计/系统开发/training_set.xlsx')
sheet = wb.active

# 训练集id保存在列表training_set
training_set = []
for row in range(2, sheet.max_row+1):
    id_cell = sheet.cell(row=row, column=1)  # 读取训练集中肿瘤id
    id_text = id_cell.value
    training_set.append(int(id_text))

# 创建数据结构 DataFrame矩阵score_dataFrame
score_dataFrame = DataFrame(columns=training_set, index=training_set)
for id in data:
    if int(id) in training_set:
        for each in data[id]:  # 确保两两计算的肿瘤均属于训练集
            if int(each) in training_set:
                score_dataFrame.loc[int(id),int(each)]=[]
                score_dataFrame.loc[int(id),int(each)].append(data[id][each]['brainStructure_score'])
                score_dataFrame.loc[int(id),int(each)].append(data[id][each]['symptoms_score'])
                score_dataFrame.loc[int(id), int(each)].append(data[id][each]['image_score'])
                score_dataFrame.loc[int(id), int(each)].append(data[id][each]['isAssociated'])

"""
ax = plt.subplot(111, projection='3d')  # 创建一个三维的绘图工程

x_data0 = []
x_data1 = []
y_data0 = []
y_data1 = []
z_data0 = []
z_data1 = []

for row in training_set:
    for col in training_set:
        if col > row:
            if score_dataFrame.loc[row,col][3] == 1:
                x_data1.append(score_dataFrame.loc[row,col][0])
                y_data1.append(score_dataFrame.loc[row,col][1])
                z_data1.append(score_dataFrame.loc[row,col][2])
            else:
                x_data0.append(score_dataFrame.loc[row,col][0])
                y_data0.append(score_dataFrame.loc[row,col][1])
                z_data0.append(score_dataFrame.loc[row,col][2])

ax.scatter(x_data0, y_data0, z_data0, c='y')  # 绘制数据点
ax.scatter(x_data1, y_data1, z_data1, c='r')

ax.set_zlabel('Z')  # 坐标轴
ax.set_ylabel('Y')
ax.set_xlabel('X')
plt.show()
"""

# 创建权重训练字典
score_dic = {}
for x in range(0,101):  # x：weight for brainStructure_score,保存为字典一级key
    score_dic[x] = {}
    for y in range(0,101-x):  # x：weight for symptoms_score,保存为字典二级key
        score_dic[x][y] = {}
        z = 100 - x - y  # x：weight for image_score,保存为字典三级key

        score = 0
        score_all = 0
        score_isAssociated = 0
        score_percent = 0

        # 根据训练集索引遍矩阵右上角历两两记录（不包括自身）
        for row in training_set:
            for col in training_set:
                if col > row:
                    score = x * score_dataFrame.loc[row, col][0] \
                            + y * score_dataFrame.loc[row, col][1] \
                            + z * score_dataFrame.loc[row, col][2]  # 单次得分记录
                    score_all += score  # 加入总得分记录
                    if score_dataFrame.loc[row, col][3] == 1:
                        score_isAssociated += score  # 为相似肿瘤，加入相似肿瘤总得分记录
        if score_all != 0:
            score_percent = score_isAssociated / score_all
        score_dic[x][y][z] = score_percent  # 字典保存相似肿瘤得分占比记录

score_max = {'x': 0, 'y': 0, 'z': 0, 'score_max': 0}
for x in score_dic:
    for y in score_dic[x]:
        for z in score_dic[x][y]:
            if score_dic[x][y][z] > score_max['score_max']:
                score_max['x'] = x/100
                score_max['y'] = y/100
                score_max['z'] = z/100
                score_max['score_max'] = score_dic[x][y][z]
print(score_max)

''''''
# 创建脑区与症状权重变化的三维的绘图工程
ax = plt.subplot(111, projection='3d')

x_data = []
y_data = []
z_data = []
for x in score_dic:
    for y in score_dic[x]:
        for z in score_dic[x][y]:
            x_data.append(x/100)
            y_data.append(y/100)
            z_data.append(score_dic[x][y][z])

ax.scatter(x_data, y_data, z_data, c='y')  # 绘制数据点

ax.set_zlabel('Score_percent')  # 坐标轴
ax.set_ylabel('Y:Weight for symptoms_score')
ax.set_xlabel('X:Weight for brainStructure_score')
# plt.show()

# 创建脑区与影像学表现权重变化的三维的绘图工程
ax = plt.subplot(111, projection='3d')

x_data = []
y_data = []
z_data = []
for x in score_dic:
    for y in score_dic[x]:
        for z in score_dic[x][y]:
            x_data.append(x/100)
            y_data.append(z/100)
            z_data.append(score_dic[x][y][z])

ax.scatter(x_data, y_data, z_data, c='y')  # 绘制数据点

ax.set_zlabel('Score_percent')  # 坐标轴
ax.set_ylabel('Z:Weight for image_score')
ax.set_xlabel('X:Weight for brainStructure_score')
# plt.show()

# 创建症状与影像学表现权重变化的三维的绘图工程
ax = plt.subplot(111, projection='3d')

x_data = []
y_data = []
z_data = []
for x in score_dic:
    for y in score_dic[x]:
        for z in score_dic[x][y]:
            x_data.append(y/100)
            y_data.append(z/100)
            z_data.append(score_dic[x][y][z])

ax.scatter(x_data, y_data, z_data, c='y')  # 绘制数据点

ax.set_zlabel('Score_percent')  # 坐标轴
ax.set_ylabel('Z:Weight for image_score')
ax.set_xlabel('Y:Weight for symptoms_score')
# plt.show()
''''''