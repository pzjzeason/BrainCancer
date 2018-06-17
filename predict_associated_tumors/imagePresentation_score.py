# -*- coding: UTF-8 -*-
import json
import warnings
warnings.filterwarnings(action='ignore', category=UserWarning, module='gensim')
from gensim import corpora, similarities, models
import numpy as np
np.set_printoptions(suppress=True) # 输出小数形式，不以科学计数法的形式输出
import re
from openpyxl import load_workbook

wb = load_workbook('/Users/燚/study/大三下/信息系统开发与设计/系统开发/data_for_python（附抽取数据）.xlsx')
sheet = wb.active  # 获得当前正在显示的sheet
sheet.insert_cols(18)  # 插入分词
sheet.insert_cols(20)


# 停用词表构建函数
def stopwordsList (filepath):
    user_stopwords = [line.strip() for line in open(filepath, 'r',encoding='utf-8').readlines()]
    return user_stopwords


stopwords = stopwordsList('/Users/燚/study/大三下/信息系统开发与设计/系统开发/stopwords.txt')  # 加载停用词路径


# 遍历excel每行文本
for row in range(2, sheet.max_row+1):
    ct_cell = sheet.cell(row=row, column=17)
    ct_text = ct_cell.value  # CT描述字符串

    mri_cell = sheet.cell(row=row, column=19)
    mri_text = ct_cell.value  # MRI描述字符串

    if ct_text:  # CT描述不为空
        ct_text = re.sub("\n", " ", str(ct_text))  # 统一格式，去除空行
        ct_text = ct_text.lower().strip()  # 统一小写

        # 提取词，其中注意一些医学描述：10-20%/小数/ T1/1mm/T1 C+ (Gd)
        ct_split = re.findall(r'[a-z]+[0-9]+|[a-z]+|[0-9]+-[0-9]+\%|[0-9]+\%|[0-9]+\.[0-9]+|[0-9]+'
                              r'|\+[0-9]+|\-[0-9]+|\–[0-9]+', ct_text)

        ct_cut = []  # 分词列表
        for word in ct_split: # 过滤停用词
            if word not in stopwords:
                ct_cut.append(word)

        sheet.cell(row=row, column=18).value = " ".join(ct_cut)

    if mri_text:
        mri_text = re.sub("\n", " ", str(mri_text))
        mri_text = mri_text.lower().strip()

        mri_split = re.findall(r'[a-z]+[0-9]+|[a-z]+\+|[a-z]+|[0-9]+-[0-9]+\%|[0-9]+\%|[0-9]+\.[0-9]+|[0-9]+|\+[0-9]+|\-[0-9]+|\–[0-9]+', ct_text)

        mri_cut = []
        for word in mri_split:
            if word not in stopwords:
                mri_cut.append(word)

        sheet.cell(row=row, column=20).value = " ".join(mri_cut)

wb.save('/Users/燚/study/大三下/信息系统开发与设计/系统开发/data_for_python（附抽取数据）.xlsx')


'''TF-IDF文本相似度计算'''
data_ct = {}
data_mri ={}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/data_for_json.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        if each.startswith(u'\ufeff'):
            each = each.encode('utf8')[3:].decode('utf8')
        line = json.loads(each)
        data_ct[line['id']] = line['ct']
        data_mri[line['id']] = line['mri']

'''制作文本库'''
corpora_documents_ct = []
corpora_documents_mri = []

for each in data_ct:
    ct_text = data_ct[each]  # 分词样本
    items_seg = []
    if isinstance(ct_text, str):  # 在成功读取一个分词序列时，将各个词写入list
        items_seg.extend(list(ct_text.split(' ')))
        corpora_documents_ct.append(items_seg)

for each in data_mri:
    # 分词样本
    mri_text = data_mri[each]
    items_seg2 = []
    if isinstance(mri_text, str):  # 在成功读取一个分词序列时，将各个词写入list
        items_seg2.extend(list(mri_text.split(' ')))
        corpora_documents_mri.append(items_seg2)

'''获取词袋（bag-of-words)，制作语料库'''
dictionary_ct = corpora.Dictionary(corpora_documents_ct)  # 生成字典和向量语料
dictionary_mri = corpora.Dictionary(corpora_documents_mri)
# dictionary_ct.save('/Users/燚/study/大三下/信息系统开发与设计/系统开发/dict.txt')  # 保存生成的词典
# dictionary_ct=dictionary_ct.load('dict.txt')#加载

dictionary_ct.keys()  # 词袋中数字作为词的索引
dictionary_mri.keys()
# print(dictionary_ct.token2id)

corpus_ct = [dictionary_ct.doc2bow(text) for text in corpora_documents_ct]  # 得到语料中每篇文档对应的稀疏向量（这里是bow向量）
corpus_mri = [dictionary_mri.doc2bow(text) for text in corpora_documents_mri]
# 向量的每一个元素代表了一个word在这篇文档中出现的次数
# print(corpus_ct)
# corpora.Mmcorpus_ct.serialize('corpus_cte.mm',corpus_ct)#保存生成的语料
# corpus_ct=corpora.Mmcorpus_ct('corpus_cte.mm')#加载

'''使用TF-IDF模型对语料库建模'''
tfidf_model_ct = models.TfidfModel(corpus_ct)  # corpus_ct是一个返回bow向量的迭代器
corpus_ct_tfidf = tfidf_model_ct[corpus_ct]  # 统计文档库中每个词（特征）的TF-IDF值
tfidf_model_mri = models.TfidfModel(corpus_mri)
corpus_ct_tfidf2 = tfidf_model_ct[corpus_mri]

# 所有文档的文本向量索引
index = similarities.SparseMatrixSimilarity(tfidf_model_ct[corpus_ct_tfidf], num_features=len(dictionary_ct.keys()))
index2 = similarities.SparseMatrixSimilarity(tfidf_model_mri[corpus_ct_tfidf2], num_features=len(dictionary_mri.keys()))

ct_score={}  # 保存为字典
mri_score={}

'''遍历文本库，对每个目标文档，分别分析其与所有文档的相似度'''
for each in data_ct:
    each_score = {}
    count = 1  # 遍历计数器
    count2 = 1
    sim = index[tfidf_model_ct[corpus_ct_tfidf[int(each)-1]]]  # 相似度计算
    sim2 = index2[tfidf_model_mri[corpus_ct_tfidf2[int(each) - 1]]]

    sim_refine=[]  # 由于相似度数值过小，进行统一*100处理
    sim2_refine=[]
    for score in sim:
        sim_refine.append(score * 100)
    for score in sim2:
        sim2_refine.append(score * 100)

    for score in sim_refine:
        each_score[str(count)] = {'ct_score': score}
        ct_score[each] = each_score
        count += 1

    score_record=ct_score[each]
    score_record["id"] = each

    for score in sim2_refine:
        score_record[str(count2)]['mri_score'] = score
        count2 += 1

    with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/image_score.json", 'a', encoding='utf-8') as json_file:
        json_file.write(json.dumps(score_record, ensure_ascii=False) + '\n')


'''影像学表现综合得分'''
data = {}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/image_score_for_compute.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        values = {}
        line = json.loads(each)

        for value in line:
            if value != 'id':
                values[value]=line[value]

        data[line['id']]=values

# 遍历每一两两相似的得分记录
for each in data:
    for value in data[each]:
        ct_score = data[each][value]['ct_score']
        mri_score = data[each][value]['mri_score']

        # 处理异常值：同一文本描述或同为空
        if ct_score > 90:
            ct_score = 0.0
        if mri_score > 90:
            mri_score=0.0

        if abs(ct_score-0)>0.01:
            if abs(mri_score-0)>0.01:
                score = (ct_score + mri_score)/2  # 若该两个肿瘤有CT和MRI两个相似，总分为平均值
            else:
                score = ct_score  # 若该两个肿瘤只有一个相似得分，取该得分为总分
        else:
            if abs(mri_score-0)>0.01:
                score = mri_score
            else:
                score = 0

        data[each][value]['score'] = score
        data[each][value].pop("ct_score")
        data[each][value].pop("mri_score")

    record = data[each]
    record['id'] = each
    with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_i.json", 'a', encoding='utf-8') as json_file:
        json_file.write(json.dumps(record, ensure_ascii=False) + '\n')