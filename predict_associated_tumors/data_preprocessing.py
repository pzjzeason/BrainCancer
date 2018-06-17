import json
from openpyxl import load_workbook
import matplotlib.pyplot as plt

'''数据归一化'''
data = {}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_with_associatedTumours.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        values = {}
        line = json.loads(each)

        for value in line:
            if value != 'id':
                values[value]=line[value]

        data[line['id']] = values

max_image_score = 0
for id in data:
    for each in data[id]:
        if data[id][each]['image_score'] > max_image_score:
            max_image_score = data[id][each]['image_score']
print(max_image_score)

max_brainStructure_score = 0
for id in data:
    for each in data[id]:
        data[id][each]['brainStructure_score'] = data[id][each]['brainStructure_score']+1
        if data[id][each]['brainStructure_score'] > max_brainStructure_score:
            max_brainStructure_score = data[id][each]['brainStructure_score']
print(max_brainStructure_score)

max_symptoms_score = 0
for id in data:
    for each in data[id]:
        data[id][each]['symptoms_score'] = data[id][each]['symptoms_score']+1
        if data[id][each]['symptoms_score'] > max_symptoms_score:
            max_symptoms_score = data[id][each]['symptoms_score']
print(max_symptoms_score)

for id in data:
    for each in data[id]:
        data[id][each]['image_score'] = data[id][each]['image_score'] / max_image_score
        data[id][each]['brainStructure_score'] = data[id][each]['brainStructure_score'] / max_brainStructure_score
        data[id][each]['symptoms_score'] = data[id][each]['symptoms_score'] / max_symptoms_score
        if  int(id) == int(each):  # 肿瘤和自身的资料为0
            data[id][each]['brainStructure_score'] = 0
            data[id][each]['symptoms_score'] = 0

for id in data:
    record = data[id]
    record['id'] = id

    with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_with_associatedTumours2.json", 'a', encoding='utf-8') as json_file:
        json_file.write(json.dumps(record, ensure_ascii=False) + '\n')


'''相似记录'''
data = {}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        values = {}
        line = json.loads(each)

        for value in line:
            if value != 'id':
                values[value]=line[value]

        data[line['id']]=values

wb = load_workbook('/Users/燚/study/大三下/信息系统开发与设计/系统开发/data_for_python.xlsx')
sheet = wb.active
for row in range(2, sheet.max_row+1):
    associatedTumour_cell = sheet.cell(row=row, column=4)  # 读取相似肿瘤记录
    associatedTumour_text = associatedTumour_cell.value

    associatedTumours = []  # 相似肿瘤列表
    if associatedTumour_text:  # 不为空
        associatedTumours.extend(list(associatedTumour_text.split(r' ')))  # 添加以空格分隔的相似肿瘤

        # 如果两两相似，则更改字典字段（对称两边均更改）
        for each in associatedTumours:
            data[str(row-1)][each]["isAssociated"] = 1
            data[each][str(row-1)]["isAssociated"] = 1

for id in data:
    for each in data[id]:
        if "isAssociated" not in data[id][each]:  #若不相似则"isAssociated"=0
            data[id][each]['isAssociated'] = 0

    record = data[id]
    record['id'] = id

    with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_with_associatedTumours.json", 'a', encoding='utf-8') as json_file:
        json_file.write(json.dumps(record, ensure_ascii=False) + '\n')