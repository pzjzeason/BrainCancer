from openpyxl import load_workbook
from openpyxl import Workbook
import re
file_txt = open('C:/Users/Thinkpad/Desktop/系统设计/structure.txt', 'r')      #将脑区结构词典导入列表中
list_structure = []
for line in file_txt.readlines():                              #按行读取文件
    line = line.strip()
    list_structure.append(line)

list_Symptoms = []
wb_Symptoms = load_workbook('C:/Users/Thinkpad/Desktop/系统设计/Symptoms.xlsx')   #将临床症状词典导入到列表中
sheet_Symptoms = wb_Symptoms.get_sheet_by_name('Sheet1')
for items in list(sheet_Symptoms.columns)[1][2:]:
    list_Symptoms.append(items.value)

WB = Workbook()            #创建一个EXCEL列表以收集数据
WB.create_sheet('Data', index=0)
sheet = WB.get_sheet_by_name('Data')
row = ['tumour', 'brainstructure', 'symptoms']
sheet.append(row)

wb = load_workbook('C:/Users/Thinkpad/Desktop/系统设计/data_for_python.xlsx')
sheet1 = wb.get_sheet_by_name('Sheet1')
for i in range(1,len(list(sheet1.columns)[1])):
    Dict = dict.fromkeys(['tumour', 'brainstructure', 'symptoms'])        #构建词典
    Dict['tumour']=list(sheet1.columns)[1][i].value                                 #脑区属性
    location = []
    string_structure=list(sheet1.columns)[22][i].value
    for model_structure in list_structure:
        if bool(re.findall(str(model_structure), str(string_structure), re.IGNORECASE)) is True:  #将结果保存到excel中
            location.append(model_structure)
    while '' in location:
        location.remove('')
    Dict['brainstructure'] = location

    symptoms=[]                                         #症状属性
    string_Symptoms=list(sheet1.columns)[14][i].value
    for model_Symptoms in list_Symptoms:
        if bool(re.findall(str(model_Symptoms), str(string_Symptoms), re.IGNORECASE)) is True:  #将结果保存到excel中
            if bool(model_structure):
                symptoms.append(model_Symptoms)
    while '' in symptoms:
        symptoms.remove('')
    Dict['symptoms']=symptoms
    excel_brainstructure = ','.join(Dict['brainstructure'])
    excel_symptoms = ','.join(Dict['symptoms'])
    row = [Dict['tumour'],  excel_brainstructure, excel_symptoms]  #写入EXCEL文件
    sheet.append(row)
WB.save(r'C:/Users/Thinkpad/Desktop/data.xlsx')                  #保存EXCEL文件




