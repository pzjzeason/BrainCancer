from owlready2 import *
from openpyxl import load_workbook
import json

# -*- coding: UTF-8 -*-
# 读取owl文件
onto = get_ontology("file:///untitled1/brainCancer.owl").load()

# 读取症状，创建症状实例
symptom_file = load_workbook('症状中英对照2.0.xlsx')
sheet_symptom = symptom_file['Sheet1']
#  循环读取excel中每一行数据，
for i in range(3, 1581):
    # 将每个单词的首字母大写，去除空格
    symptom_en = sheet_symptom.cell(i, 2).value
    symptom_cn = sheet_symptom.cell(i, 1).value
    splitName1 = symptom_en.split(sep=' ')
    splitName2 = []
    for each in splitName1:
        splitName2.append(each.title())
    finalName = ''.join(splitName2)
    s = onto.Symptoms(finalName, namespace=onto)
    s.NameInCN.append(symptom_cn)

# 读取易感人群数据，创建易感人群实例
sheet_people = load_workbook('易患人群中英对照.xlsx')['Sheet1']
for i in range(1, 7):
    # 读取易患人群的英文和中文形式，并去除字符串前后的空格
    peopleName_en = (sheet_people.cell(i, 1).value).strip(' ')
    peopleName_cn = (sheet_people.cell(i, 2).value).strip(' ')
    # 在本体中创建易患人群的实例
    people = onto.People(peopleName_en,namespace=onto)
    people.NameInCN.append(peopleName_cn)

# 读取脑区数据，创建脑区实例
sheet_brainStructure = load_workbook('脑区中英文对照.xlsx')['Sheet1']
for i in range(2, 50):
    # 读取脑区的英文和中文形式，并去除字符串前后的空格
    brainStructure_en = (sheet_brainStructure.cell(i, 5).value).strip(' ')
    brainStructure_cn = (sheet_brainStructure.cell(i, 6).value).strip(' ')
    brainStructure_parent = ((sheet_brainStructure.cell(i,7).value).strip(' '))

    # 去除字符串中间的空格，并将空格后的单词首字母大写(第一个单词首字母小写)
    splitString1_brain = brainStructure_en.split(sep=' ')
    number = len(splitString1_brain)
    for each in range(0, number):
        if each == 0:
            splitString1_brain[each] = splitString1_brain[each].lower()
    finalBrainStructure = ''.join(splitString1_brain)
    # 父类字符串处理
    splitString1_brainParent = brainStructure_parent.split(sep=' ')
    splitString2_brainParent = []
    for each in splitString1_brainParent:
        splitString2_brainParent.append(each[0].upper() + each[1:])
    finalBrainStructureParent = ''.join(splitString2_brainParent)
    # 在本体中创建脑区的实例
    brainStructure = onto[finalBrainStructureParent](name = finalBrainStructure)
    brainStructure.NameInCN.append(brainStructure_cn)

# 读取检查项目数据，创建检查项目实例
sheet_diagnosis = load_workbook('检查项目.xlsx')['Sheet1']
for i in range(2, 20):
    # 读取易患人群的英文和中文形式，并去除字符串前后的空格
    diagnosis_en = (sheet_diagnosis.cell(i, 1).value).strip(' ')
    diagnosis_cn = (sheet_diagnosis.cell(i, 2).value).strip(' ')
    diagnosis_other = sheet_diagnosis.cell(i, 3).value
    splitDiagnosis = diagnosis_en.split(sep=' ')
    splitDiagnosis2 = []
    for each in splitDiagnosis:
        splitDiagnosis2.append(each[0].upper() + each[1:])
    finalDiagnosisName = ''.join(splitDiagnosis2)
    # 在本体中创建检查项目的实例
    diagnosis = onto.Diagnosis(finalDiagnosisName, namespace=onto)
    splitInCn = diagnosis_cn.split(';')
    for name in splitInCn:

        diagnosis.NameInCN.append(name)
    if diagnosis_other is not None:
        eachReferences = diagnosis_other.split(';')
        for each in eachReferences:
            diagnosis.DiagnosisOtherName.append(locstr(each, lang="en"))



# 读取肿瘤数据表格,创建肿瘤实例
tumour_file = load_workbook('data_for_python2.xlsx')
sheet_tumour = tumour_file['Sheet1']

# 列表，元素为每一个肿瘤
t = []

# 循环读取excel中每一行的数据
for i in range(1, 166):
    # 读取肿瘤的英文名称，并去除字符串前后的空格
    name1 = sheet_tumour.cell(i+1, 2).value
    name2 = name1.strip(' ')

    # 去除字符串中间的空格，并将空格后的单词首字母大写
    splitName1 = name2.split(sep=' ')
    splitName2 = []
    for each in splitName1:
        splitName2.append(each[0].upper() + each[1:])
    finalName = ''.join(splitName2)

    # 读取肿瘤的大类名称，同样去除字符串中间的空格，并将空格后的单词首字母大写
    parent = sheet_tumour.cell(i + 1, 10).value
    splitName3 = parent.split(sep=' ')
    splitName4 = []
    for each in splitName3:
        splitName4.append(each[0].upper() + each[1:])
    finalParentName = ''.join(splitName4)

    # 将肿瘤名称填充到owl中, 英文名字作为该肿瘤的IRI名字后缀
    t.append(onto[finalParentName](name = finalName))

    # 填充注释属性和数据属性
    tumourNameCn = sheet_tumour.cell(i+1, 1).value
    name_also_en = sheet_tumour.cell(i+1, 4).value
    name_also_cn = sheet_tumour.cell(i+1, 3).value
    icd10_code = sheet_tumour.cell(i+1, 5).value
    icd_thematicWord = sheet_tumour.cell(i+1, 6).value
    icd_O_code = sheet_tumour.cell(i+1, 7).value
    meshCode = sheet_tumour.cell(i+1, 8).value
    mesh_thematicWord = sheet_tumour.cell(i+1, 9).value
    clinicDepartment = sheet_tumour.cell(i+1, 11).value
    definition = sheet_tumour.cell(i+1, 12).value
    isDefinedBy = sheet_tumour.cell(i+1, 13).value
    cause = sheet_tumour.cell(i+1, 14).value
    CT_description = sheet_tumour.cell(i+1, 16).value
    MRI_description = sheet_tumour.cell(i+1, 17).value
    images = sheet_tumour.cell(i+1, 18).value
    DifferentialDiagnosis = sheet_tumour.cell(i+1, 19).value
    references = sheet_tumour.cell(i+1, 24).value
    incidence = sheet_tumour.cell(i+1, 25).value
    associatedOther = sheet_tumour.cell(i + 1, 27).value
    definitionInCn = sheet_tumour.cell(i+1, 28).value
    causeInCn = sheet_tumour.cell(i+1, 29).value
    DifferentialDiagnosisInCn = sheet_tumour.cell(i+1, 30).value
    referencesInCn = sheet_tumour.cell(i+1, 31).value
    CT_descriptionInCn = sheet_tumour.cell(i+1, 32).value
    MRI_descriptionInCn = sheet_tumour.cell(i+1, 33).value

    if name_also_cn is not None:
        eachName_also_cn = name_also_cn.split(';')
        for each in eachName_also_cn:
            t[i - 1].nameAlso.append(locstr(each, lang="cn"))
    if name_also_en is not None:
        eachName_also_en = name_also_en.split(';')
        for each in eachName_also_en:
            t[i - 1].nameAlso.append(locstr(each, lang="en"))
    if definition is not None:
        t[i-1].definition.append(locstr(definition, lang="en"))
    if isDefinedBy is not None:
        t[i-1].isDefinedBy = isDefinedBy
    if references is not None:
        eachReferences = references.split('\n')
        for each in eachReferences:
            t[i-1].references.append(locstr(each, lang="en"))
    if tumourNameCn is not None:
        t[i-1].NameInCN.append(tumourNameCn)
    if icd10_code is not None:
        t[i-1].ICD_10Code = icd10_code
    if icd_thematicWord is not None:
        t[i-1].ICDThematicWord = icd_thematicWord
    if icd_O_code is not None:
        t[i-1].ICD_OCode = icd_O_code
    if meshCode is not None:
        t[i-1].MeshCode = meshCode
    if DifferentialDiagnosis is not None:
        t[i-1].DifferentialDiagnosis = DifferentialDiagnosis
    if mesh_thematicWord is not None:
        t[i-1].MeshThematicWord = mesh_thematicWord
    if clinicDepartment is not None:
        t[i-1].ClinicDepartment = clinicDepartment
    if cause is not None:
        t[i-1].Pathogeny = cause
    if CT_description is not None:
        t[i-1].CTImageDescription = CT_description
    if MRI_description is not None:
        t[i-1].MRIImageDescription = MRI_description
    if incidence is not None:
        t[i-1].Incidence = incidence
    if associatedOther is not None:
        t[i-1].OtherAssociatedIllness = associatedOther
    if images is not None:
        image_Diction = json.loads(images)
        image_number = list(image_Diction.keys())
        t[i-1].Image = image_number
    if definitionInCn is not None:
        t[i - 1].definition.append(locstr(definitionInCn, lang="cn"))
    if causeInCn is not None:
        t[i-1].PathogenyInCN = causeInCn
    if DifferentialDiagnosisInCn is not None:
        t[i - 1].DifferentialDiagnosisInCN = DifferentialDiagnosisInCn
    if referencesInCn is not None:
        eachReferences = referencesInCn.split('\n')
        for each in eachReferences:
            t[i - 1].references.append(locstr(each, lang="cn"))
    if CT_descriptionInCn is not None:
        t[i - 1].CTImageDescriptionInCN = CT_descriptionInCn
    if MRI_descriptionInCn is not None:
        t[i - 1].MRIImageDescriptionInCN = MRI_descriptionInCn
    # 如果该肿瘤为斜体，为其添加斜体注释属性
    italicTumors = ['DiffuseAstrocytoma,IDH-wildtype', 'AnaplasticAstrocytoma,IDH-wildtype', 'EpithelioidGlioblastoma',
                        'AnaplasticOligodendroglioma,NOS','Oligoastrocytoma,NOS',
                    'AnaplasticOligodendroglioma', 'DiffuseLeptomeningealGlioneuronalTumour',
                        'Medulloblastomas,group3','Medulloblastomas,group4','EmbryonalTumorsWithMultilayeredRosettes,C20MC-altered'
                    ,'CNSEmbryonalTumorWithRhabdoidFeatures']
    if finalName in italicTumors:
        t[i - 1].italic = '在2016WHO中枢神经系统肿瘤分类中，该肿瘤以斜体字表示，即为暂定的肿瘤实体，尚不能将其确定为独立疾病'


# 填充对象属性
    # 建立肿瘤实例和临床表现实例之间的肿瘤临床表现（hasSymptoms）关系
    symptom = sheet_tumour.cell(i + 1, 15).value
    # 如果症状不为空
    if symptom is not None:
        # 用','将字符串隔开，统一症状的格式，找到实例，在肿瘤实例和症状实例之间建立hasSymptoms对象属性
        splitName_symptom = symptom.split(sep=',')
        for each in splitName_symptom:
            splitName1_symptom = each.split(sep=' ')
            splitName2_symptom = []
            for each2 in splitName1_symptom:
                splitName2_symptom.append(each2.title())
            finalName_symptom = ''.join(splitName2_symptom)
            tSymptom = onto[finalName_symptom]
            t[i - 1].hasSymptoms.append(tSymptom)


    #建立肿瘤实例和人群实例之间的肿瘤易发人群（highOccurrenceRateGroup）关系
    # 读取肿瘤的易患人群
    highOccurrenceRateGroup = sheet_tumour.cell(i+1, 22).value
    # 如果易患人群不为空
    if highOccurrenceRateGroup is not None:
        # 以“，”为分隔符，将易患人群的字符串隔开，并在相应的肿瘤实例和人群实例之间建立关系
        splitGroup = (highOccurrenceRateGroup.strip(' ')).split(sep=',')
        for each in splitGroup:
            tPeople = onto[each]
            t[i-1].hasHighOccurrenceRateIn.append(tPeople)


     #建立肿瘤实例和脑区实例之间的肿瘤易发脑区（findingSite）关系
     # 读取肿瘤的易发脑区
    findingSite = sheet_tumour.cell(i + 1, 23).value
    if findingSite is not None:
            # 用','将字符串隔开，统一脑区的格式
            splitString_site = findingSite.split(sep=',')
            for each in splitString_site:

                if each != 'skull' and each != '非脑区':
                    splitString1_site = each.split(sep=' ')
                    finalFindingSite = ''.join(splitString1_site)

                    # 在相应的肿瘤实例和脑区实例之间建立hasFindingSite关系
                    tBrainStructure = onto[finalFindingSite]

                    # 判断tBrainStructure是类还是实例
                    if type(tBrainStructure) == type(onto.BrainStructure):
                        # 判断这个类有没有下一级的类
                        second = tBrainStructure.subclasses()
                        # 如果有下一级的类，再判断是否还有下一级，如果还有下一级就继续判断，直到没有，将t[i-1]与类的实例构建关系
                        if second != []:
                            for sub in second:
                                third = sub.subclasses()
                                if third != []:
                                    for th in third:
                                        individuals = th.instances()
                                        for indi in individuals:
                                            t[i - 1].hasFindingSite.append(indi)
                                else:
                                    individuals = list(sub.instances())
                                    for indi in individuals:
                                        t[i - 1].hasFindingSite.append(indi)
                        # 如果下一级为空，就将t[i-1]与它的实例构建关系
                        else:
                            individuals = list(tBrainStructure.instances())
                            for indi in individuals:
                                t[i - 1].hasFindingSite.append(indi)
                    # 如果是实例，在t[i-1]与实例间构建关系
                    else:
                        t[i - 1].hasFindingSite.append(tBrainStructure)


    #建立肿瘤实例和检查项目实例之间的肿瘤检查项目（diagnosis）关系
    diagnosis = sheet_tumour.cell(i + 1, 21).value
    # 读取检查项目
    if diagnosis is not None:
        diagnosis = diagnosis.strip(' ')
        splitName_diagnosis = diagnosis.split(sep='、')
        for each in splitName_diagnosis:
            splitName1_diagnosis = each.split(sep=' ')
            splitName2_diagnosis = []
            for each2 in splitName1_diagnosis:
                splitName2_diagnosis.append(each2[0].upper() + each2[1:])
            finalName_diagnosis = ''.join(splitName2_diagnosis)
            tDiagnosis = onto[finalName_diagnosis]
            t[i - 1].isDiagnosedBy.append(tDiagnosis)


#建立肿瘤之间相似病（associatedBrain）关系
for i in range(1, 166):
    # 读取文献资料中的相似病
    associatedBrain = sheet_tumour.cell(i+1, 26).value

    # 相似是CentralNervousSystemTumours
    if associatedBrain is not None:
        # 如果字符串最后一个字符是‘;’,就将最后一个字符删掉
        if associatedBrain[-1] == ';':
            associatedBrain = associatedBrain[:-1]
        # 通过分号将字符串隔开
        split_associatedBrain = associatedBrain.split(sep=';')
        for each in split_associatedBrain:
            splitName1_associatedBrain = (each.strip(' ')).split(sep=' ')
            splitName2_associatedBrain = []
            for each2 in splitName1_associatedBrain:
                splitName2_associatedBrain.append(each2[0].upper() + each2[1:])
            finalName_associatedBrain = ''.join(splitName2_associatedBrain)
            tAssociated = onto[finalName_associatedBrain]
            t[i - 1].associatedTumours.append(tAssociated)

            tAssociated.associatedTumours.append(t[i - 1])

# 填充计算预测的相似病
sheet_ProbablyAssociated = load_workbook('相似肿瘤预测（计算）.xlsx')['Sheet1']
for i in range(1, 11):
    # 读取肿瘤及相似肿瘤，并将字符串格式规范化
    brainTumour = (sheet_ProbablyAssociated.cell(i, 1).value).strip(' ')
    splitName1 = brainTumour.split(sep=' ')
    splitName2 = []
    for each in splitName1:
        splitName2.append(each[0].upper() + each[1:])
    brainTumour = ''.join(splitName2)

    probablyAssociated = (sheet_ProbablyAssociated.cell(i, 2).value).strip(' ')
    splitName1Associated = probablyAssociated.split(sep=' ')
    splitName2Associated = []
    for each in splitName1Associated:
        splitName2Associated.append(each[0].upper() + each[1:])
    probablyAssociated = ''.join(splitName2Associated)
    # 构建对象属性
    tumour = onto[brainTumour]
    associatedWith = onto[probablyAssociated]
    tumour.associatedTumours.append(associatedWith)
    # 加上注释，说明这一对相似肿瘤是通过计算预测出来的
    comment[tumour, onto.associatedTumours, associatedWith] = '此相似肿瘤是通过计算预测出的'

# 填充推理预测的相似病
sheet_ProbablyAssociated = load_workbook('相似肿瘤预测（推理）.xlsx')['Sheet1']
for i in range(1, 93):
        # 读取肿瘤及相似肿瘤，并将字符串格式规范化
        brainTumour = (sheet_ProbablyAssociated.cell(i, 1).value).strip(' ')
        splitName1 = brainTumour.split(sep=' ')
        splitName2 = []
        for each in splitName1:
            splitName2.append(each[0].upper() + each[1:])
        brainTumour = ''.join(splitName2)
        probablyAssociated = (sheet_ProbablyAssociated.cell(i, 2).value).strip(' ')
        splitName1Associated = probablyAssociated.split(sep=' ')
        splitName2Associated = []
        for each in splitName1Associated:
            splitName2Associated.append(each[0].upper() + each[1:])
        probablyAssociated = ''.join(splitName2Associated)
        # 构建对象属性
        tumour = onto[brainTumour]
        associatedWith = onto[probablyAssociated]
        tumour.associatedTumours.append(associatedWith)
        # 加上注释，说明这一对相似肿瘤是通过计算预测出来的
        comment[tumour, onto.associatedTumours, associatedWith] = '此相似肿瘤是通过推理预测出的'


# 构建肿瘤之间的上下级关系
sheet_belong = load_workbook('子类父类.xlsx')['Sheet1']
for i in range(2, 27):
    # 读取肿瘤及其父类，并将其格式规范化
    children = sheet_belong.cell(i, 1).value
    children = children.strip(' ')
    splitName1 = children.split(sep=' ')
    splitName2 = []
    for each in splitName1:
        splitName2.append(each[0].upper() + each[1:])
    children = ''.join(splitName2)
    parent = sheet_belong.cell(i, 2).value
    parent = parent.strip(' ')
    splitName1_Parent = parent.split(sep=' ')
    splitName2_Parent = []
    for each in splitName1_Parent:
        splitName2_Parent.append(each[0].upper() + each[1:])
    parent = ''.join(splitName2_Parent)
    # 构建对象属性
    t_Children = onto[children]
    t_Parent = onto[parent]
    t_Children.isChildOf.append(t_Parent)


# 存储,括号内为路径
filename = 'D:/untitled1/braincancer3.owl'
onto.save(filename)
