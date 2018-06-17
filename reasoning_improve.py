from pyecharts import Graph
import json
import matplotlib.pyplot as plt
from openpyxl import load_workbook

similar_dic = {}
with open("/Users/燚/study/大三下/信息系统开发与设计/系统开发/score/score_with_associatedTumours2.json", "r", encoding='utf-8') as json_file:
    for each in json_file.readlines():
        line = json.loads(each)
        similar_dic[int(line['id'])] = []

        for value in line:
            if value != 'id':
                if line[value]['isAssociated'] == 1:
                    similar_dic[int(line['id'])].append(int(value))

nodes = []
for each in similar_dic:
    each_node = {}
    if len(similar_dic[each])!=0:
        each_node['name'] = str(each)
        each_node['symbolSize'] = 10+2*len(similar_dic[each])
        nodes.append(each_node)
print(nodes)

links = []
similar_list=[]
for i in similar_dic:
    if len(similar_dic[i]) != 0:
        for j in similar_dic[i]:
            if j>i:
                similar_list.append([i,j])


# a-b-c
predict1_list=[]
predict1_dic={}
for i in range(1,33):
    predict1_dic[i]=[]
for i in similar_dic:
    if len(similar_dic[i])!=0:
        for j in similar_dic[i]:
            for k in similar_dic[i]:
                if k != j:
                    each_predict=[]
                    each_predict=[min(k,j),max(k,j)]
                    if each_predict not in predict1_list:
                        predict1_list.append(each_predict)
                        for item in range(1,33):
                            if len(similar_dic[k])>=item or len(similar_dic[j])>=item:
                                predict1_dic[item].append(each_predict)
count=0
for each in predict1_list:
    if each in similar_list:
        count+=1
        links.append({"source": str(each[0]), "target": str(each[1])})
    else:
        links.append({"source": str(each[0]), "target": str(each[1]),
                      "lineStyle": {"normal": {"width": 0.5,"type": "dotted","color": "teal"}}})

print(count/len(predict1_list))

"""
p_data = []
r_data = []

for each in predict1_dic:
    count = 0
    r_data.append(each)
    for item in predict1_dic[each]:
        if item in similar_list:
            count+=1
    p_data.append(count/len(predict1_dic[each]))

plt.plot(r_data, p_data, 'o-')  # 绘制数据点
ax = plt.gca()

# 绘制x，y轴的说明
plt.xlabel('Degree Centrality')
plt.ylabel('Reasoning Precision')
plt.title('Degree-Reasoning Curve')
plt.show()  # 显示

# a-b-c-d
predict2_list=[]
for i in similar_dic:
    if len(similar_dic[i])!=0:
        for j in similar_dic[i]:
            if len(similar_dic[j])!=0:
                for m in similar_dic[j]:
                    for k in similar_dic[i]:
                        if k!=j:
                            each_predict = []
                            each_predict = [min(k, m), max(k, m)]
                            if each_predict not in predict2_list:
                                predict2_list.append(each_predict)
count=0
for each in predict2_list:
    if each in similar_list:
        count+=1
print(count/len(predict2_list))

# a-b-c-d-e
predict3_list = []
for i in similar_dic:
    if len(similar_dic[i])!=0:
        for j in similar_dic[i]:
            if len(similar_dic[j])!=0:
                for m in similar_dic[j]:
                    if len(similar_dic[m])!=0:
                        for n in similar_dic[m]:
                            for k in similar_dic[i]:
                                if k != j:
                                    each_predict = []
                                    each_predict = [min(k, n), max(k, n)]
                                    if each_predict not in predict3_list:
                                        predict3_list.append(each_predict)
count=0
for each in predict3_list:
    if each in similar_list:
        count+=1
print(count/len(predict3_list))

graph = Graph("肿瘤显著相似关系图",width=1200, height=600)
graph.add("", nodes,links,graph_repulsion=800,is_label_show=True)
graph.render()

wb = load_workbook('/Users/燚/study/大三下/信息系统开发与设计/系统开发/prediction2.xlsx')
sheet = wb.active  # 获得当前正在显示的sheet

count = 0  # 预测相似计数器
for each in predict1_dic[26]:
    if each not in similar_list:
        row_id = each[0]
        col_id = each[1]
        sheet.cell(row=count+1, column=1).value = row_id
        sheet.cell(row=count+1, column=2).value = col_id
        count+=1
wb.save('/Users/燚/study/大三下/信息系统开发与设计/系统开发/prediction2.xlsx')
"""
graph={}
graph['name']=nodes
graph['link']=links
with open(r"\Users\燚\PycharmProjects\brainCancer\graph.json", 'a',encoding='utf-8') as json_file:
    json_file.write(json.dumps(graph, ensure_ascii=False) + '\n')